#!/usr/bin/env python

import functions
import glob
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter
from tkinter import messagebox, filedialog

# du könntest diese ganzen Parameter auch als Kommandozeilen Parameter übergebbar machen.
# schau dir dazu mal argparse an: https://docs.python.org/3/library/argparse.html
today = datetime.now()
# snapshots older than 'age' are considered old.
age = 30
# outlook mails are encoded as 'latin-1'
encoding = 'latin-1'
# table headers as formated in the emails)
vmTableHeaders = ['VMname', 'VMOwner', 'SSName', 'SSCreated', 'SSDescription']
# du kannst dir einmal das Mapping merken von Spaltenname zu Index
headerIndices = dict((v, i) for i, v in enumerate(vmTableHeaders))
# headers for the output excel table
excelHeaders = ['ServerName', '', 'ServerOwner', 'Date', 'Description']
# du kannst neue Listen auch einfach initialisieren als
oldSnapshots = []
filesWithoutHTML = []

# ask user for folder with emails (as html)
root = tkinter.Tk()
root.withdraw()
tkinter.messagebox.showinfo('Folder Select', 'Please select the folder that contains the emails (saved as html-Files)')
filePath = filedialog.askdirectory() + '/'

filesInFolder = glob.glob(os.path.join(filePath, '*.htm'))
filesInFolder.extend(glob.glob(os.path.join(filePath, '*.html')))

# Ich habe mir angewöhnt, den ganzen Code, der wirklich irgendwas macht, in Python Modulen normalerweise in ein
# if __name__ == '__main__':
# zu packen. das stellt sicher, dass der Code nur ausgeführt wird, wenn du die Datei direkt ausführst und nicht,
# wenn du sie nur in einem anderen Python Modul importierst. Das würde hier wahrscheinlich keinen Unterschied machen
# ist aber meiner Meinung nach guter Stil.

if not filesInFolder:
    tkinter.messagebox.showinfo \
        ('Parsing aborted', 'The selected folder doesn\'t seem to contain any htm- or html-files')
    quit()

for file in filesInFolder:
    emailTables = functions.getTablesFromHTML(file, encoding)
    vmTable = []
    # check if one of the tables is the one with the correct header
    for table in emailTables:
        if table[0] == vmTableHeaders:
            vmTable = table
    if not vmTable:
        filesWithoutHTML.append(os.path.realpath(file))
        continue

    # test all rows except the first if date is older than 'age' and vm-name starts with 'z'
    for row in table[1:]:
        # aus Gründen der Übersichtlichkeit könntest du die Elemente, die du dir aus der row ziehst und
        # auf die du vergleichst, nochmal temporären Variablen mit sprechenden Namen zuweisen
        if row[headerIndices['VMname']].startswith('z') and (today - timedelta(days=age)) > \
                datetime.strptime(row[headerIndices['SSCreated']], '%m/%d/%Y %I:%M:%S %p'):
            # du könntest sogar die ganze Zeile in ein schönes Dict umwandeln
            # row = dict((vmTableHeaders[i], v) for i, v in enumerate(row))
            # dann kannst du später überall direkt snapshot['VMname'] etc verwenden
            oldSnapshots.append(row)

if not oldSnapshots:
    tkinter.messagebox.showinfo \
        ('Parsing aborted', 'None of the files in \"' + filePath + '\" contains a table of VM snapshots.')
    # benutz stattdessen besser sys.exit()
    quit()

# arbeite hier lieber mit dicts
from collections import defaultdict
serverOwners = defaultdict(list)
for snapshot in oldSnapshots:
    serverOwners[snapshot[vmTableHeaders.index('VMOwner')]].append(snapshot)

# creating output xlsx-file
book = Workbook()
sheet = book.active
sheet.append(excelHeaders)
# append snapshot info to sheet, group them by owner
for owner in sorted(serverOwners):
    oldSnapshots = serverOwners[owner]
    for snapshot in oldSnapshots:
        # je nach dem, was du von meinen Vorschlägen oben verwendet hast, kanns du den Zugriff auf die Spaltenwerte
        # der Zeile hier vereinfachen
        line = [snapshot[vmTableHeaders.index('VMname')].rstrip(),
                '-',
                snapshot[vmTableHeaders.index('VMOwner')].rstrip(),
                datetime.strptime(snapshot[vmTableHeaders.index('SSCreated')].split(' ')[0], '%m/%d/%Y').strftime(
                    '%Y-%m-%d').rstrip(),
                snapshot[vmTableHeaders.index('SSDescription')].rstrip()]
        sheet.append(line)

# formatting: first row bold and bigger; all column widths adjusted to maximum length of content
for cell in sheet["1:1"]:
    cell.font = Font(size=16, bold=True)
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 1

# add the filenames where no tables of snapshots were found
rowPointer = sheet.max_row + 1
if filesWithoutHTML:
    filesWithoutHTML.insert(0, 'Files in Folder without a table of VM snapshots:')
    for entry in filesWithoutHTML:
        rowPointer += 1
        sheet.cell(rowPointer, 1, entry).font = Font(color='FF0000')

errorString = 'Not executed yet.'
saveFile = filePath + 'alteSnapshots-' + today.strftime('%Y-%m-%d') + '.xlsx'
# die Whileschleife am Error String laufen zu lassen, finde ich seltsam. da würde ich eher ein writeSuccess = False nehmen
# und das in der Schleife dann im Erfolgsfall auf True setzen, oder so
while errorString:
    try:
        book.save(saveFile)
        os.startfile(os.path.realpath(filePath))
        errorString = None
    except OSError as e:
        ok = tkinter.messagebox.askokcancel \
            ('Write Error', 'The file \"' + saveFile + '\" is open.\n' \
             'Close it and press OK to overwrite it or press cancel to abort the program.')
        if ok:
            # in case I want to do something with the error string later
            errorString = e
        else:
            errorString = None
